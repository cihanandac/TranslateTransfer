import pandas as pd
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl import load_workbook

#pd.options.mode.chained_assignment = None
root = tk.Tk()
root.withdraw()

# Asks users to choose the both excel files from their system.
filepath = filedialog.askopenfilename(title="Choose your excel file")
translations_path = filedialog.askopenfilename(title="Choose translation file")

file = pd.ExcelFile(filepath)
sheets = file.sheet_names

translations = pd.ExcelFile(translations_path)
tsheets = translations.sheet_names


wb = load_workbook(filepath)


for sheet in sheets:
    print("Now working on "+sheet+" sheet")
    page = file.parse(sheet)
    lenght, widht = page.shape

    # ws is our worksheet that the work on in this iteration
    ws = wb[sheet]
    for tsheet in tsheets:
        print("Now you are working on " + tsheet + " translation sheet")
        tpage = translations.parse(tsheet)
        tlenght, twidth = tpage.shape

        # Both 'Inv. No.' and 'ObjectNumber' is the name of the columns at each excel file.
        for i in range(1, lenght):
            for k in range(1, tlenght):
                a = page['Inv. No.'][i]
                b = tpage['ObjectNumber'][k]

                # Again 'Description_EN' and 'Remarks' are the name of the columns which contains translation text.
                if a == b:
                    print("Found a match and will transfer the translation.")
                    ws['AC'+str(i+2)] = tpage['Description_EN'][k]
                    ws['AD'+str(i+2)] = tpage['Remarks'][k]

wb.save(filepath)
